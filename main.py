import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import os
import logging
import shutil
import threading
import queue
import json
import re
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache

# 3rd‑party deps
import dateutil.parser  # pip install python-dateutil
from docx import Document  # pip install python-docx
from openpyxl import load_workbook  # pip install openpyxl

# local deps
from text_extraction import extract_text_from_pdf, extract_text_from_image

# -----------------------------
# Configuration & Logging
# -----------------------------
CONFIG_FILE = "config.json"
COMPANIES_FILE = "firmen.txt"
LOG_FILE = "app.log"

DEFAULT_CONFIG = {
    "DEFAULT_SOURCE_DIR": "",
    "BACKUP_DIR": "backup",
    "ALLOWED_EXTENSIONS": ["pdf", "png", "jpg", "jpeg", "docx", "xlsx", "eml"],
    "BATCH_SIZE": 10,
    "DATE_FORMATS": ["%Y.%m.%d", "%Y-%m-%d", "%d.%m.%Y"],
    "MAIN_TARGET_DIR": "",
    "LOG_LEVEL": "INFO",
    "FILENAME_PATTERN": "{date}_{company}_RE-{number}.{ext}",
    "DARK_MODE": False,
    "STRIP_LEGAL_SUFFIXES": True,
    "LEGAL_SUFFIXES": [
        "GmbH & Co. KG", "GmbH & Co KG", "GmbH & Co.KG",
        "UG (haftungsbeschränkt)", "UG", "GmbH", "AG", "KG", "OHG", "GbR", "e.K.", "e.V."
    ]
}


class ConfigManager:
    def __init__(self, path: str = CONFIG_FILE):
        self.path = path
        self.data = DEFAULT_CONFIG.copy()
        self.load()

    def load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                # Only allow known keys; fall back to defaults if missing
                self.data.update({k: raw.get(k, v) for k, v in DEFAULT_CONFIG.items()})
            except Exception as e:
                logging.error("Konfiguration fehlerhaft – es werden Standardwerte verwendet: %s", e)
        self.apply_logging()

    def save(self):
        tmp = self.path + ".tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            os.replace(tmp, self.path)
        except Exception as e:
            logging.error("Konfiguration konnte nicht gespeichert werden: %s", e)
            if os.path.exists(tmp):
                os.remove(tmp)

    def apply_logging(self):
        # Configure logging ONCE, replacing existing handlers cleanly
        root = logging.getLogger()
        for h in list(root.handlers):
            root.removeHandler(h)
        level = getattr(logging, self.data.get("LOG_LEVEL", "INFO").upper(), logging.INFO)
        logging.basicConfig(
            level=level,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
            handlers=[
                logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8"),
                logging.StreamHandler()
            ],
        )


config = ConfigManager()

# -----------------------------
# Date parsing
# -----------------------------
GERMAN_MONTHS = {
    "Januar": "January", "Februar": "February", "März": "March", "April": "April",
    "Mai": "May", "Juni": "June", "Juli": "July", "August": "August",
    "September": "September", "Oktober": "October", "November": "November", "Dezember": "December",
    "Jan": "Jan", "Feb": "Feb", "Mär": "Mar", "Mar": "Mar", "Apr": "Apr", "Mai": "May",
    "May": "May", "Jun": "Jun", "Jul": "Jul", "Aug": "Aug", "Sep": "Sep", "Okt": "Oct",
    "Oct": "Oct", "Nov": "Nov", "Dez": "Dec", "Dec": "Dec",
}

DATE_PATTERNS = [
    (r"\b(\d{4})[-/.](\d{2})[-/.](\d{2})\b", "%Y-%m-%d"),  # 2024-02-03
    (r"\b(\d{2})[-/.](\d{2})[-/.](\d{4})\b", "%d.%m.%Y"),  # 03.02.2024
    (r"\b(\d{2})/(\d{2})/(\d{4})\b", "%m/%d/%Y"),          # 02/03/2024
    (r"\b(\d{4})/(\d{2})/(\d{2})\b", "%Y/%m/%d"),          # 2024/02/03
    (r"\b(\d{1,2})[.\s]?(Jan|Feb|Mär|Mar|Apr|Mai|May|Jun|Jul|Aug|Sep|Okt|Oct|Nov|Dez|Dec)[.]?\s?(\d{4})\b", "%d %b %Y"),
    (r"\b(\d{1,2}) (January|February|March|April|May|June|July|August|September|October|November|December) (\d{4})\b", "%d %B %Y"),
    (r"\b(\d{8})\b", "%Y%m%d"),
    (r"\b(\d{1,2}) (Januar|Februar|März|April|Mai|Juni|Juli|August|September|Oktober|November|Dezember) (\d{4})\b", "%d %B %Y"),
]


def detect_date(text: str) -> str | None:
    if not text:
        return None
    # normalize German month names to English for strptime
    norm = text
    for ger, eng in GERMAN_MONTHS.items():
        norm = re.sub(rf"\b{re.escape(ger)}\b", eng, norm, flags=re.IGNORECASE)
    # try explicit formats first
    for pattern, fmt in DATE_PATTERNS:
        m = re.search(pattern, norm, re.IGNORECASE)
        if m:
            try:
                val = " ".join(m.groups())
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    # fallback to dateutil
    try:
        return dateutil.parser.parse(norm, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return None


# -----------------------------
# Plugins
# -----------------------------
class Plugin:
    extensions: tuple[str, ...] = ()

    def process_file(self, path: str) -> str:
        raise NotImplementedError


class PDFPlugin((Plugin)):
    extensions = ("pdf",)

    def process_file(self, path: str) -> str:
        return extract_text_from_pdf(path)


class ImagePlugin(Plugin):
    extensions = ("png", "jpg", "jpeg")

    def process_file(self, path: str) -> str:
        return extract_text_from_image(path)


class EMLPlugin(Plugin):
    extensions = ("eml",)

    def process_file(self, path: str) -> str:
        import email
        from email import policy
        from email.parser import BytesParser
        with open(path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)
        body = msg.get_body(preferencelist=("plain",))
        return body.get_content() if body else ""


class WordPlugin(Plugin):
    extensions = ("docx",)

    def process_file(self, path: str) -> str:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)


class ExcelPlugin(Plugin):
    extensions = ("xlsx",)

    def process_file(self, path: str) -> str:
        wb = load_workbook(path, data_only=True, read_only=True)
        out: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                out.append(" ".join(str(c) for c in row if c is not None))
        return "\n".join(out)


PLUGINS: list[Plugin] = [PDFPlugin(), ImagePlugin(), EMLPlugin(), WordPlugin(), ExcelPlugin()]
EXT_TO_PLUGIN: dict[str, Plugin] = {
    ext: plugin for plugin in PLUGINS for ext in plugin.extensions
}


@lru_cache(maxsize=512)
def extract_text_cached(path: str) -> str:
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    plugin = EXT_TO_PLUGIN.get(ext)
    if not plugin:
        logging.warning("Nicht unterstützte Erweiterung: %s", ext)
        return ""
    try:
        return plugin.process_file(path)
    except Exception as e:
        logging.error("Textextraktion fehlgeschlagen (%s): %s", path, e)
        return ""


# -----------------------------
# Text analysis
# -----------------------------
@dataclass
class ExtractedInfo:
    company_name: str = "Unbekannt"
    date: str = ""
    number: str = ""


def _strip_legal_suffix(name: str) -> str:
    if not name:
        return name
    if not config.data.get("STRIP_LEGAL_SUFFIXES", True):
        return name
    suffixes = sorted(config.data.get("LEGAL_SUFFIXES", []), key=len, reverse=True)
    pattern = r"\s+(?:" + "|".join(map(re.escape, suffixes)) + r")\.?$"
    return re.sub(pattern, "", name, flags=re.IGNORECASE).strip()


@lru_cache(maxsize=1024)
def analyze_text(text: str, known_companies: tuple[str, ...] = ()) -> ExtractedInfo:
    info = ExtractedInfo()
    try:
        # 1) Firma (Liste > Heuristik)
        for c in sorted(known_companies, key=len, reverse=True):
            if re.search(rf"\b{re.escape(c)}\b", text):
                info.company_name = c
                break
        else:
            company_keywords = ["GmbH", "GbR", "OHG", "AG", "KG", "UG", "e.K.", "e.V.", "GmbH & Co. KG"]
            m = re.search(rf"([\wÄÖÜäöüß&.,\- ]+)\s+(?:{'|'.join(map(re.escape, company_keywords))})", text)
            if m:
                info.company_name = m.group(1).strip()

        info.company_name = _strip_legal_suffix(info.company_name)

        # 2) Datum (explizite Muster > Fallback)
        date_matches = re.findall(r"\b\d{2}\.\d{2}\.\d{4}\b|\b\d{4}-\d{2}-\d{2}\b|\b\d{8}\b", text)
        for d in date_matches:
            detected = detect_date(d)
            if detected:
                info.date = detected
                break
        if not info.date:
            detected = detect_date(text)
            if detected:
                info.date = detected

        # 3) Rechnungsnummer (de/en + Varianten)
        token = r"([A-Za-z0-9][A-Za-z0-9._\-/#]{3,})"
        patterns = [
            rf"Rechnung\s*(?:Nr|No|#|Nummer|ID)?\.?\s*[:\-#]?\s*{token}",
            rf"Rechn\.?\s*(?:Nr|No)?\.?\s*[:\-#]?\s*{token}",
            rf"Rg\.?\s*Nr\.?\s*[:\-#]?\s*{token}",
            rf"Beleg(?:nr|nummer)?\s*[:\-#]?\s*{token}",
            rf"Invoice\s*(?:No|#|ID)?\s*[:\-#]?\s*{token}",
            rf"Document\s*(?:No|#|ID)?\s*[:\-#]?\s*{token}",
            rf"Dokumentnummer\s*[:\-#]?\s*{token}",
        ]
        for p in patterns:
            m = re.search(p, text, flags=re.IGNORECASE)
            if m:
                info.number = m.group(1).strip()
                if info.number.startswith("AEU"):
                    info.company_name = "Amazon"
                break
        # Fallback: typische SAP-IDs (10–12 Ziffern)
        if not info.number:
            m = re.search(r"\b\d{10,12}\b", text)
            if m:
                info.number = m.group(0)
    except Exception as e:
        logging.error("Analyse fehlgeschlagen: %s", e)
    return info


def format_filename(info: ExtractedInfo, ext: str) -> str:
    safe_company = re.sub(r"[\\/:*?\"<>|]", "_", info.company_name).strip() or "Unbekannt"
    safe_number  = re.sub(r"[\\/:*?\"<>|\s]", "_", info.number)
    date = info.date or "0000-00-00"
    return config.data["FILENAME_PATTERN"].format(
        date=date, company=safe_company, number=safe_number, ext=ext
    )


# -----------------------------
# File operations
# -----------------------------
@dataclass
class ProcessResult:
    src: str
    dst: str | None
    error: str | None


def ensure_backup(path: str):
    try:
        backup_dir = os.path.join(os.path.dirname(path), config.data["BACKUP_DIR"])
        os.makedirs(backup_dir, exist_ok=True)
        shutil.copy2(path, backup_dir)
    except Exception as e:
        logging.warning("Backup fehlgeschlagen (%s): %s", path, e)


def move_unique(src: str, dst: str) -> str:
    base, ext = os.path.splitext(dst)
    candidate = dst
    n = 1
    while os.path.exists(candidate):
        candidate = f"{base}_{n}{ext}"
        n += 1
    shutil.move(src, candidate)
    return candidate


def process_one_file(path: str, known_companies: tuple[str, ...]) -> ProcessResult:
    try:
        ensure_backup(path)
        text = extract_text_cached(path)
        info = analyze_text(text, known_companies)
        ext = os.path.splitext(path)[1].lstrip(".").lower()

        # target folder selection
        year = (info.date or "0000-00-00").split("-")[0]
        root_target = config.data.get("MAIN_TARGET_DIR") or os.path.dirname(path)
        year_dir = os.path.join(root_target, year)
        company_base = _strip_legal_suffix(info.company_name)
        company_dir = os.path.join(year_dir, re.sub(r"[\\/:*?\"<>|]", "_", company_base) or "Unbekannt")
        os.makedirs(company_dir, exist_ok=True)

        new_name = format_filename(info, ext)
        dst = os.path.join(company_dir, new_name)
        dst = move_unique(path, dst)
        return ProcessResult(src=path, dst=dst, error=None)
    except Exception as e:
        try:
            err_dir = os.path.join(os.path.dirname(path), "errors")
            os.makedirs(err_dir, exist_ok=True)
            shutil.move(path, os.path.join(err_dir, os.path.basename(path)))
        except Exception:
            pass
        return ProcessResult(src=path, dst=None, error=str(e))


# -----------------------------
# GUI Utilities
# -----------------------------
class Tooltip:
    def __init__(self, widget: tk.Widget, text: str):
        self.widget = widget
        self.text = text
        self.tipwindow: tk.Toplevel | None = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, event=None):
        if self.tipwindow:
            return
        x = self.widget.winfo_rootx() + 10
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify=tk.LEFT, relief=tk.SOLID, borderwidth=1,
                         background="#ffffe0")
        label.pack(ipadx=4, ipady=2)

    def hide(self, event=None):
        if self.tipwindow is not None:
            self.tipwindow.destroy()
            self.tipwindow = None


LANG = {
    'de': {
        'source_label': 'Quellverzeichnis:',
        'button_source': 'Quellverzeichnis auswählen',
        'button_rename': 'Dateien umbenennen und organisieren',
        'button_preview': 'Vorschau der Dateibenennung',
        'button_firmenpflege': 'Firmenpflege',
        'button_help': 'Hilfe',
        'button_report': 'Bericht anzeigen',
        'button_exit': 'Beenden',
        'button_config': 'Konfiguration',
        'button_log': 'Protokoll anzeigen',
        'button_info': 'Info',
        'tooltip_source': 'Wählen Sie das Verzeichnis mit den Dateien.',
        'tooltip_rename': 'Startet die Umbenennung im Hintergrund.',
        'tooltip_preview': 'Zeigt die neuen Dateinamen an.',
        'tooltip_firmenpflege': 'Bearbeiten Sie die Firmenliste.',
        'tooltip_config': 'Passen Sie die Einstellungen an.',
        'tooltip_report': 'Zeigt einen Verarbeitungsbericht.',
        'tooltip_log': 'Öffnet die Log-Datei.',
        'tooltip_help': 'Zeigt diese Hilfe.',
        'tooltip_exit': 'Programm beenden.',
        'tooltip_info': 'Informationen zum Tool.',
        'tooltip_language': 'Sprache wählen.'
    },
    'en': {
        'source_label': 'Source Directory:',
        'button_source': 'Select Source Directory',
        'button_rename': 'Rename and Organize Files',
        'button_preview': 'Preview Filenames',
        'button_firmenpflege': 'Manage Companies',
        'button_help': 'Help',
        'button_report': 'Show Report',
        'button_exit': 'Exit',
        'button_config': 'Configuration',
        'button_log': 'Show Log',
        'button_info': 'Info',
        'tooltip_source': 'Choose the folder with files.',
        'tooltip_rename': 'Start background processing.',
        'tooltip_preview': 'Preview new filenames.',
        'tooltip_firmenpflege': 'Edit company list.',
        'tooltip_config': 'Adjust settings.',
        'tooltip_report': 'Show processing report.',
        'tooltip_log': 'Open the log file.',
        'tooltip_help': 'Show this help.',
        'tooltip_exit': 'Quit program.',
        'tooltip_info': 'About this tool.',
        'tooltip_language': 'Choose UI language.'
    }
}


# -----------------------------
# App
# -----------------------------
class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Dateiumbenennungstool")
        self.lang = tk.StringVar(value='de')
        self.source_dir = tk.StringVar(value=config.data["DEFAULT_SOURCE_DIR"]) 

        # queues for worker communication
        self.task_q: queue.Queue[str] = queue.Queue()
        self.result_q: queue.Queue[ProcessResult] = queue.Queue()
        self.worker: threading.Thread | None = None
        self.running = False

        # report storage
        self.processed: list[ProcessResult] = []

        self._build_ui()
        self._apply_style()
        self._load_companies()
        self.root.after(200, self._poll_results)

    def _build_ui(self):
        l = LANG[self.lang.get()]
        # Frames
        self.frame_dirs = tk.LabelFrame(self.root, text="Verzeichnisse", padx=10, pady=10)
        self.frame_dirs.grid(row=0, column=0, padx=10, pady=10, sticky='ew')
        self.root.columnconfigure(0, weight=1)

        # Source chooser
        self.lbl_source = tk.Label(self.frame_dirs, text=l['source_label'])
        self.lbl_source.grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.ent_source = tk.Entry(self.frame_dirs, textvariable=self.source_dir, width=60)
        self.ent_source.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        self.btn_source = tk.Button(self.frame_dirs, text=l['button_source'], command=self._select_dir)
        self.btn_source.grid(row=0, column=2, padx=5, pady=5)
        Tooltip(self.btn_source, l['tooltip_source'])

        self.btn_run = tk.Button(self.frame_dirs, text=l['button_rename'], command=self._start_processing)
        self.btn_run.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky='ew')
        Tooltip(self.btn_run, l['tooltip_rename'])

        self.btn_preview = tk.Button(self.frame_dirs, text=l['button_preview'], command=self._preview)
        self.btn_preview.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky='ew')
        Tooltip(self.btn_preview, l['tooltip_preview'])

        # Config frame
        self.frame_cfg = tk.LabelFrame(self.root, text="Konfiguration", padx=10, pady=10)
        self.frame_cfg.grid(row=1, column=0, padx=10, pady=10, sticky='ew')

        self.btn_companies = tk.Button(self.frame_cfg, text=l['button_firmenpflege'], command=self._edit_companies)
        self.btn_companies.grid(row=0, column=0, padx=5, pady=5)

        self.btn_config = tk.Button(self.frame_cfg, text=l['button_config'], command=self._open_config)
        self.btn_config.grid(row=0, column=1, padx=5, pady=5)

        # Reports frame
        self.frame_reports = tk.LabelFrame(self.root, text="Berichte", padx=10, pady=10)
        self.frame_reports.grid(row=2, column=0, padx=10, pady=10, sticky='ew')

        self.btn_report = tk.Button(self.frame_reports, text=l['button_report'], command=self._show_report)
        self.btn_report.grid(row=0, column=0, padx=5, pady=5)

        self.btn_log = tk.Button(self.frame_reports, text=l['button_log'], command=self._show_log)
        self.btn_log.grid(row=0, column=1, padx=5, pady=5)

        # Misc frame
        self.frame_misc = tk.LabelFrame(self.root, text="Sonstige", padx=10, pady=10)
        self.frame_misc.grid(row=3, column=0, padx=10, pady=10, sticky='ew')

        self.btn_help = tk.Button(self.frame_misc, text=l['button_help'], command=self._show_help)
        self.btn_help.grid(row=0, column=0, padx=5, pady=5)

        self.btn_exit = tk.Button(self.frame_misc, text=l['button_exit'], command=self.root.quit)
        self.btn_exit.grid(row=0, column=1, padx=5, pady=5)

        self.btn_info = tk.Button(self.frame_misc, text=l['button_info'], command=self._show_info)
        self.btn_info.grid(row=0, column=2, padx=5, pady=5)

        # Progress & file list
        self.progress = ttk.Progressbar(self.root, orient="horizontal", mode="determinate")
        self.progress.grid(row=4, column=0, padx=10, pady=10, sticky='ew')

        self.files_list = tk.Listbox(self.root, width=80, height=10)
        self.files_list.grid(row=5, column=0, padx=10, pady=10, sticky='nsew')
        self.root.rowconfigure(5, weight=1)

        # Language chooser
        self.cmb_lang = ttk.Combobox(self.root, textvariable=self.lang, values=list(LANG.keys()), state="readonly")
        self.cmb_lang.grid(row=6, column=0, padx=10, pady=10, sticky='e')
        self.cmb_lang.bind("<<ComboboxSelected>>", self._change_language)

    def _apply_style(self):
        if config.data.get("DARK_MODE"):
            try:
                self.root.configure(bg="#111")
                for child in self.root.winfo_children():
                    self._apply_dark_recursive(child)
            except Exception as e:
                logging.warning("Dark Mode konnte nicht vollständig angewendet werden: %s", e)

    def _apply_dark_recursive(self, widget):
        try:
            if isinstance(widget, (tk.Frame, tk.LabelFrame)):
                widget.configure(bg="#111", fg="#eee")
            elif isinstance(widget, tk.Label):
                widget.configure(bg="#111", fg="#eee")
            elif isinstance(widget, tk.Button):
                widget.configure(bg="#222", fg="#eee", activebackground="#333")
        except Exception:
            pass
        for child in widget.winfo_children():
            self._apply_dark_recursive(child)

    def _load_companies(self):
        if os.path.exists(COMPANIES_FILE):
            try:
                with open(COMPANIES_FILE, "r", encoding="utf-8-sig") as f:
                    self.companies = tuple(line.strip() for line in f if line.strip())
            except Exception as e:
                logging.error("Firmenliste konnte nicht geladen werden: %s", e)
                self.companies = tuple()
        else:
            self.companies = tuple()

    # ---------- Actions ----------
    def _select_dir(self):
        d = filedialog.askdirectory()
        if d and os.path.exists(d):
            self.source_dir.set(d)

    def _start_processing(self):
        directory = self.source_dir.get()
        if not directory:
            messagebox.showwarning("Warnung", "Quellverzeichnis nicht ausgewählt.")
            return
        # gather files
        allowed = set(ext.lower() for ext in config.data["ALLOWED_EXTENSIONS"])
        files = [os.path.join(directory, f) for f in os.listdir(directory) if f.split(".")[-1].lower() in allowed]
        if not files:
            messagebox.showwarning("Warnung", "Keine verarbeitbaren Dateien gefunden.")
            return
        self.files_list.delete(0, tk.END)
        for f in files:
            self.files_list.insert(tk.END, os.path.basename(f))
        self.progress.configure(maximum=len(files), value=0)

        # start worker thread
        self.running = True
        for f in files:
            self.task_q.put(f)
        if not self.worker or not self.worker.is_alive():
            self.worker = threading.Thread(target=self._worker_loop, daemon=True)
            self.worker.start()

    def _worker_loop(self):
        while self.running and not self.task_q.empty():
            path = self.task_q.get()
            res = process_one_file(path, self.companies)
            self.result_q.put(res)
        self.running = False

    def _poll_results(self):
        updated = False
        while True:
            try:
                res: ProcessResult = self.result_q.get_nowait()
            except queue.Empty:
                break
            else:
                updated = True
                if res.error:
                    logging.error("Fehler bei %s: %s", res.src, res.error)
                else:
                    logging.info("Verschoben: %s -> %s", res.src, res.dst)
                self.processed.append(res)
                self.progress.step(1)
        if updated and self.progress['value'] >= self.progress['maximum']:
            messagebox.showinfo("Erfolg", "Alle Dateien verarbeitet.")
        self.root.after(200, self._poll_results)

    def _preview(self):
        directory = self.source_dir.get()
        if not directory:
            messagebox.showwarning("Warnung", "Quellverzeichnis nicht ausgewählt.")
            return
        allowed = set(ext.lower() for ext in config.data["ALLOWED_EXTENSIONS"])
        files = [f for f in os.listdir(directory) if f.split(".")[-1].lower() in allowed]
        if not files:
            messagebox.showwarning("Warnung", "Keine Dateien gefunden.")
            return
        win = tk.Toplevel(self.root)
        win.title("Vorschau der Dateibenennung")
        lb = tk.Listbox(win, width=100, height=25)
        lb.pack(fill=tk.BOTH, expand=True)
        for fn in files:
            path = os.path.join(directory, fn)
            text = extract_text_cached(path)
            info = analyze_text(text, self.companies)
            ext = fn.split(".")[-1]
            new_name = format_filename(info, ext)
            lb.insert(tk.END, f"{fn} -> {new_name}")
        tk.Button(win, text="Schließen", command=win.destroy).pack(pady=6)

    def _edit_companies(self):
        win = tk.Toplevel(self.root)
        win.title("Firmenpflege")
        lb = tk.Listbox(win)
        lb.pack(fill=tk.BOTH, expand=True)
        for c in self.companies:
            lb.insert(tk.END, c)

        def add():
            name = simpledialog.askstring("Neue Firma", "Bitte Firmennamen eingeben:", parent=win)
            if name:
                lb.insert(tk.END, name)

        def remove():
            sel = lb.curselection()
            if sel:
                lb.delete(sel[0])

        def save():
            items = [lb.get(i) for i in range(lb.size())]
            try:
                with open(COMPANIES_FILE, "w", encoding="utf-8") as f:
                    f.write("\n".join(items))
                self._load_companies()
                messagebox.showinfo("Gespeichert", "Firmenliste aktualisiert.")
            except Exception as e:
                messagebox.showerror("Fehler", f"Konnte Firmenliste nicht speichern: {e}")

        tk.Button(win, text="Hinzufügen", command=add).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(win, text="Entfernen", command=remove).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(win, text="Speichern", command=save).pack(side=tk.RIGHT, padx=5, pady=5)

    def _open_config(self):
        win = tk.Toplevel(self.root)
        win.title("Konfiguration")
        entries = {}

        def add_entry(row, label, key):
            tk.Label(win, text=label).grid(row=row, column=0, sticky='w', padx=6, pady=4)
            e = tk.Entry(win, width=60)
            e.grid(row=row, column=1, sticky='ew', padx=6, pady=4)
            e.insert(0, str(config.data.get(key, "")))
            entries[key] = e

        add_entry(0, "Standard-Quellverzeichnis:", "DEFAULT_SOURCE_DIR")
        add_entry(1, "Backup-Verzeichnis:", "BACKUP_DIR")
        add_entry(2, "Erlaubte Erweiterungen (Komma):", "ALLOWED_EXTENSIONS")
        add_entry(3, "Batch-Größe:", "BATCH_SIZE")
        add_entry(4, "Datumsformate (Komma):", "DATE_FORMATS")
        add_entry(5, "Hauptzielordner:", "MAIN_TARGET_DIR")
        add_entry(6, "Dateinamenmuster:", "FILENAME_PATTERN")

        tk.Label(win, text="Log-Level:").grid(row=7, column=0, sticky='w', padx=6, pady=4)
        cb = ttk.Combobox(win, values=["DEBUG", "INFO", "WARNING", "ERROR"], state="readonly")
        cb.set(config.data.get("LOG_LEVEL", "INFO"))
        cb.grid(row=7, column=1, sticky='w', padx=6, pady=4)

        dark_var = tk.BooleanVar(value=config.data.get("DARK_MODE", False))
        tk.Checkbutton(win, text="Dark Mode", variable=dark_var).grid(row=8, column=1, sticky='w', padx=6, pady=4)

        def save_changes():
            try:
                config.data["DEFAULT_SOURCE_DIR"] = entries["DEFAULT_SOURCE_DIR"].get()
                config.data["BACKUP_DIR"] = entries["BACKUP_DIR"].get()
                config.data["ALLOWED_EXTENSIONS"] = [x.strip().lower() for x in entries["ALLOWED_EXTENSIONS"].get().split(',') if x.strip()]
                config.data["BATCH_SIZE"] = int(entries["BATCH_SIZE"].get())
                config.data["DATE_FORMATS"] = [x.strip() for x in entries["DATE_FORMATS"].get().split(',') if x.strip()]
                config.data["MAIN_TARGET_DIR"] = entries["MAIN_TARGET_DIR"].get()
                config.data["FILENAME_PATTERN"] = entries["FILENAME_PATTERN"].get()
                config.data["LOG_LEVEL"] = cb.get()
                config.data["DARK_MODE"] = bool(dark_var.get())
                config.save()
                config.apply_logging()
                self._apply_style()
                win.destroy()
            except Exception as e:
                messagebox.showerror("Fehler", f"Speichern fehlgeschlagen: {e}")

        tk.Button(win, text="Speichern", command=save_changes).grid(row=9, column=0, columnspan=2, pady=8)

    def _show_report(self):
        end = datetime.now()
        report_lines = [
            "Bericht über die Dateiverarbeitung",
            "==============================",
            f"Datum und Uhrzeit: {end.strftime('%Y-%m-%d %H:%M:%S')}",
            f"Anzahl verarbeitet: {sum(1 for r in self.processed if r.dst)}",
            f"Fehler: {sum(1 for r in self.processed if r.error)}",
            "",
            "Details:",
        ]
        for r in self.processed:
            if r.error:
                report_lines.append(f"FEHLER: {r.src} -> {r.error}")
            else:
                report_lines.append(f"OK: {r.src} -> {r.dst}")
        text = "\n".join(report_lines)

        win = tk.Toplevel(self.root)
        win.title("Bericht")
        t = tk.Text(win, wrap='word')
        t.insert('1.0', text)
        t.config(state='disabled')
        t.pack(expand=True, fill='both')
        def copy():
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.root.update()
        tk.Button(win, text="In Zwischenablage kopieren", command=copy).pack(pady=6)

    def _show_log(self):
        try:
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
        except FileNotFoundError:
            messagebox.showerror("Fehler", f"{LOG_FILE} nicht gefunden.")
            return
        win = tk.Toplevel(self.root)
        win.title("Protokoll")
        t = tk.Text(win, wrap='word')
        t.insert('1.0', content)
        t.config(state='disabled')
        t.pack(expand=True, fill='both')

    def _show_help(self):
        messagebox.showinfo("Hilfe", (
            "1) Quellverzeichnis wählen.\n"
            "2) 'Dateien umbenennen und organisieren' starten.\n"
            "3) Vorschau nutzen, um Dateinamen zu prüfen.\n"
            "4) Einstellungen (Zielordner, Muster) anpassen.\n"
            "5) Bericht/Log für Details öffnen."
        ))

    def _show_info(self):
        try:
            with open('toolinfo.json', 'r', encoding='utf-8') as f:
                info = json.load(f)
            text = (f"Name: {info.get('name','')}\n"
                    f"Version: {info.get('version','')}\n"
                    f"Autor: {info.get('author','')}\n"
                    f"Beschreibung: {info.get('description','')}\n"
                    f"Homepage: {info.get('homepage','')}\n"
                    f"Kategorien: {', '.join(info.get('categories', []))}\n"
                    f"Features: {', '.join(info.get('features', []))}\n"
                    f"Abhängigkeiten: {', '.join(info.get('dependencies', []))}\n"
                    f"Installation: {info.get('installation',{}).get('requirements','')}\n"
                    f"Verwendung: {info.get('installation',{}).get('usage','')}")
            messagebox.showinfo("Info", text)
        except Exception as e:
            messagebox.showerror("Fehler", f"toolinfo.json Problem: {e}")

    def _change_language(self, _evt=None):
        l = LANG.get(self.lang.get(), LANG['de'])
        self.lbl_source.config(text=l['source_label'])
        self.btn_source.config(text=l['button_source'])
        self.btn_run.config(text=l['button_rename'])
        self.btn_preview.config(text=l['button_preview'])
        self.btn_companies.config(text=l['button_firmenpflege'])
        self.btn_config.config(text=l['button_config'])
        self.btn_report.config(text=l['button_report'])
        self.btn_log.config(text=l['button_log'])
        self.btn_help.config(text=l['button_help'])
        self.btn_exit.config(text=l['button_exit'])
        self.btn_info.config(text=l['button_info'])


# -----------------------------
# Main entry
# -----------------------------

def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
