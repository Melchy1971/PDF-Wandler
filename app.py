import os
import re
import json
import shutil
import threading
import queue
import subprocess
import sys
import hashlib
import csv
from pathlib import Path
from typing import Optional, Callable, List, Tuple, Dict, Any

import pdfplumber
from pdf2image import convert_from_path
from PIL import Image
import pytesseract
import dateparser
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ===============================================================
# Konfiguration / Config-Datei
# ===============================================================
APP_NAME = "Rechnungs-Assistent"
CONFIG_PATH = Path("config.json")
DEFAULT_CONFIG = {
    "input_dir": "input",
    "output_dir": "output",
    "unknown_dirname": "unbekannt",
    "max_pages_ocr": 3,
    "langs_ocr": "deu+eng",
    "use_ollama": False,
    "ollama_model": "llama3",
    "ollama_url": "http://localhost:11434/api/generate",
    # Exporte & Prüfen
    "export_csv": True,
    "export_xlsx": True,
    "review_before_save": True
}

COMPANY_KEYWORDS = [
    "GmbH", "AG", "KG", "UG", "OHG", "e.K.", "GmbH & Co. KG", "Inc", "LLC", "Limited", "Ltd", "eG"
]

RE_DATE_HINTS = re.compile(
    r"(rechnungsdatum|ausgestellt\s*am|datum)\s*[:\-]?\s*(?P<date>\d{1,2}\.\d{1,2}\.\d{2,4}|\d{4}-\d{1,2}-\d{1,2})",
    re.IGNORECASE
)
RE_DATE_FALLBACK = re.compile(r"(?P<date>\d{1,2}\.\d{1,2}\.\d{2,4}|\d{4}-\d{1,2}-\d{1,2})")
RE_INVOICE_NO = re.compile(
    r"(rechnungs(?:nummer|nr\.? )|rechnung\s*#?|invoice\s*(no\.|#)?)\s*[:\-]?\s*(?P<no>[A-Z0-9\-\/]{4,})",
    re.IGNORECASE
)

# ----------------- Config-Helper -----------------

def load_config():
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception:
            cfg = DEFAULT_CONFIG.copy()
    else:
        cfg = DEFAULT_CONFIG.copy()
        save_config(cfg)
    for k, v in DEFAULT_CONFIG.items():
        cfg.setdefault(k, v)
    return cfg


def save_config(cfg: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# ===============================================================
# Extraktion & PDF-Handling
# ===============================================================

def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def extract_text_from_pdf(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texts = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                if t.strip():
                    texts.append(t)
            return "\n".join(texts)
    except Exception:
        return ""


def ocr_pdf(pdf_path: Path, max_pages: int, langs_ocr: str) -> str:
    text_chunks = []
    try:
        images = convert_from_path(str(pdf_path), first_page=1, last_page=max_pages, dpi=300)
        for im in images:
            txt = pytesseract.image_to_string(im, lang=langs_ocr)
            if txt.strip():
                text_chunks.append(txt)
    except Exception:
        pass
    return "\n".join(text_chunks)


def guess_supplier(full_text: str) -> Optional[str]:
    lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]
    head = lines[:15]
    for ln in head:
        for kw in COMPANY_KEYWORDS:
            if kw.lower() in ln.lower():
                cand = re.split(r"[•\|,;/]{1}", ln)[0].strip()
                return cand[:80]
    before_invoice = []
    for ln in head:
        if re.search(r"\brechnung\b", ln, re.IGNORECASE):
            break
        before_invoice.append(ln)
    if before_invoice:
        cand = before_invoice[0]
        return re.sub(r"\s{2,}", " ", cand)[:80]
    return None


def parse_date(s: Optional[str]):
    if not s:
        return None
    dt = dateparser.parse(s, settings={"DATE_ORDER": "DMY", "PREFER_DAY_OF_MONTH": "first"})
    return dt.date() if dt else None


def parse_invoice_fields(full_text: str) -> dict:
    data = {"supplier": None, "invoice_no": None, "invoice_date": None}

    m = RE_DATE_HINTS.search(full_text)
    if m:
        data["invoice_date"] = parse_date(m.group("date"))
    else:
        m2 = RE_DATE_FALLBACK.search(full_text)
        if m2:
            data["invoice_date"] = parse_date(m2.group("date"))

    m3 = RE_INVOICE_NO.search(full_text)
    if m3:
        raw = m3.group("no")
        raw = re.sub(r"[^\w\-\/]", "", raw)
        data["invoice_no"] = raw[:40]

    data["supplier"] = guess_supplier(full_text)
    return data


def normalize_supplier_with_ollama(supplier: str, cfg: dict) -> str:
    if not supplier or not cfg.get("use_ollama"):
        return supplier
    try:
        import requests
        prompt = (
            "Extrahiere den reinen Firmennamen aus diesem Text, ohne Adresse oder Zusätze. "
            "Gib nur den Namen zurück:\n\n" + supplier
        )
        payload = {"model": cfg.get("ollama_model", "llama3"), "prompt": prompt, "stream": False}
        r = requests.post(cfg.get("ollama_url", "http://localhost:11434/api/generate"), json=payload, timeout=10)
        name = r.json().get("response", "").strip()
        name = re.sub(r"[\n\r\t]", " ", name)
        name = re.sub(r"\s{2,}", " ", name)
        return name[:80] if name else supplier
    except Exception:
        return supplier


def first_page_as_image(pdf_path: Path):
    try:
        images = convert_from_path(str(pdf_path), first_page=1, last_page=1, dpi=200)
        return images[0] if images else None
    except Exception:
        return None


def write_summary_pdf(dst_path: Path, fields: dict, source_pdf: Path, preview_image: Optional[Image.Image]):
    dst_path.parent.mkdir(parents=True, exist_ok=True)
    c = rl_canvas.Canvas(str(dst_path), pagesize=A4)
    width, height = A4
    margin = 20 * mm
    y = height - margin

    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, "Rechnungs-Zusammenfassung")
    y -= 12 * mm

    def line(label, value):
        nonlocal y
        c.setFont("Helvetica", 11)
        c.drawString(margin, y, f"{label}: {value}")
        y -= 8 * mm

    dateval = fields.get("invoice_date").isoformat() if fields.get("invoice_date") else "—"
    line("Rechnungsdatum", dateval)
    line("Lieferant", fields.get("supplier") or "—")
    line("Rechnungsnummer", fields.get("invoice_no") or "—")
    line("Quelle", str(source_pdf.name))

    y -= 5 * mm
    c.line(margin, y, width - margin, y)
    y -= 5 * mm

    if preview_image is not None:
        max_w = width - 2*margin
        target_h = 120 * mm
        im_w, im_h = preview_image.size
        scale = min(max_w / im_w, target_h / im_h)
        new_size = (int(im_w * scale), int(im_h * scale))
        thumb = preview_image.resize(new_size)
        c.drawImage(ImageReader(thumb), margin, y - new_size[1], width=new_size[0], height=new_size[1], preserveAspectRatio=True, mask='auto')
        y -= new_size[1] + 5 * mm

    c.showPage()
    c.save()


def build_base_name(fields: dict, unknown_dirname: str, cfg: dict) -> Tuple[str, str, str]:
    date_part = fields.get("invoice_date")
    supplier = fields.get("supplier")
    inv_no = fields.get("invoice_no")

    if supplier:
        supplier = normalize_supplier_with_ollama(supplier, cfg)
        supplier = re.sub(r"[\\/:*?\"<>|]", "_", supplier)

    if date_part:
        year = str(date_part.year)
        date_str = date_part.isoformat()
    else:
        year = unknown_dirname
        date_str = "unknown-date"

    if not supplier:
        supplier = unknown_dirname
    if not inv_no:
        inv_no = "unknown-no"

    base_name = f"{date_str}_{supplier}_Re-{inv_no}"
    return year, supplier, base_name


# ===============================================================
# CSV / Excel Export
# ===============================================================

def open_csv_writer(output_dir: Path):
    f = open(output_dir / "export.csv", "w", newline="", encoding="utf-8")
    fieldnames = ["Dateiname", "Neuer Name", "Lieferant", "Rechnungsnummer", "Datum", "Hash", "Zielpfad"]
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    return f, writer


def open_xlsx_book(output_dir: Path):
    if not OPENPYXL_AVAILABLE:
        return None, None
    path = output_dir / "export.xlsx"
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Dateiname", "Neuer Name", "Lieferant", "Rechnungsnummer", "Datum", "Hash", "Zielpfad"])
    return wb, ws


def append_row_exports(csv_writer, ws, row: List[Any]):
    if csv_writer:
        csv_writer.writerow({
            'Dateiname': row[0], 'Neuer Name': row[1], 'Lieferant': row[2],
            'Rechnungsnummer': row[3], 'Datum': row[4], 'Hash': row[5], 'Zielpfad': row[6]
        })
    if ws is not None:
        ws.append(row)


# ===============================================================
# Verarbeitung mit optionalem Review-Callback aus der GUI
# ===============================================================

def process_pdf(
    pdf_path: Path,
    cfg: dict,
    log: Callable[[str], None] = lambda msg: None,
    csv_writer=None,
    ws=None,
    seen_hashes: Optional[set] = None,
    review_cb: Optional[Callable[[dict, Path], Optional[dict]]] = None,
):
    unknown_dirname = cfg.get("unknown_dirname", "unbekannt")
    output_dir = Path(cfg.get("output_dir", "output"))

    log(f"Verarbeite: {pdf_path.name}")

    # Duplikatprüfung
    filehash = file_hash(pdf_path)
    if seen_hashes is not None and filehash in seen_hashes:
        log(f"→ Duplikat erkannt, übersprungen: {pdf_path.name}")
        return
    if seen_hashes is not None:
        seen_hashes.add(filehash)

    # Textextraktion + Felder
    text = extract_text_from_pdf(pdf_path)
    if not text.strip():
        text = ocr_pdf(pdf_path, cfg.get("max_pages_ocr", 3), cfg.get("langs_ocr", "deu+eng"))

    fields = parse_invoice_fields(text)

    # Review-Dialog (optional)
    if cfg.get("review_before_save") and review_cb is not None:
        edited = review_cb(fields, pdf_path)
        if edited is None:
            log("→ Abgebrochen durch Nutzer.")
            return
        fields = edited

    # Fallback: unbekannt
    if not any([fields.get("invoice_no"), fields.get("supplier"), fields.get("invoice_date")]):
        target_folder = output_dir / unknown_dirname
        target_folder.mkdir(parents=True, exist_ok=True)
        dst = target_folder / pdf_path.name
        shutil.copy2(pdf_path, dst)
        log(f"→ Keine Felder erkannt. Verschoben nach: {dst}")
        return

    year, supplier, base_name = build_base_name(fields, unknown_dirname, cfg)
    folder = output_dir / year / supplier
    folder.mkdir(parents=True, exist_ok=True)

    # 1) Original unter neuem Namen
    dst_original = folder / f"{base_name}.pdf"
    shutil.copy2(pdf_path, dst_original)

    # 2) Zusammenfassung
    preview = first_page_as_image(pdf_path)
    dst_summary = folder / f"{base_name}-summary.pdf"
    write_summary_pdf(dst_summary, fields, pdf_path, preview_image=preview)

    log(f"→ Original gespeichert: {dst_original}")
    log(f"→ Zusammenfassung gespeichert: {dst_summary}")

    date_str = fields.get("invoice_date").isoformat() if fields.get("invoice_date") else ""
    append_row_exports(
        csv_writer,
        ws,
        [
            pdf_path.name,
            dst_original.name,
            fields.get("supplier") or "",
            fields.get("invoice_no") or "",
            date_str,
            filehash,
            str(dst_original)
        ],
    )


def process_all(
    cfg: dict,
    log: Callable[[str], None] = lambda msg: None,
    selected_files: Optional[List[str]] = None,
    progress: Optional[Callable[[int, int], None]] = None,
    review_cb: Optional[Callable[[dict, Path], Optional[dict]]] = None,
):
    input_dir = Path(cfg.get("input_dir", "input"))
    output_dir = Path(cfg.get("output_dir", "output"))
    input_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    pdfs = [Path(f) for f in selected_files] if selected_files else list(input_dir.glob("*.pdf"))
    total = len(pdfs)
    if not pdfs:
        log("Keine PDF-Dateien gefunden.")
        if progress:
            progress(0, 1)
        return

    seen_hashes: set = set()

    csv_file = None
    csv_writer = None
    if cfg.get("export_csv"):
        csv_file, csv_writer = open_csv_writer(output_dir)

    wb = None
    ws = None
    if cfg.get("export_xlsx") and OPENPYXL_AVAILABLE:
        wb, ws = open_xlsx_book(output_dir)

    for idx, p in enumerate(pdfs, 1):
        try:
            process_pdf(p, cfg, log=log, csv_writer=csv_writer, ws=ws, seen_hashes=seen_hashes, review_cb=review_cb)
        except Exception as e:
            unk = output_dir / cfg.get("unknown_dirname", "unbekannt")
            unk.mkdir(parents=True, exist_ok=True)
            shutil.copy2(p, unk / p.name)
            log(f"Fehler bei {p.name}: {e}\n→ Datei nach '{unk}' kopiert.")
        finally:
            if progress:
                progress(idx, total)

    if csv_file:
        csv_file.close()
    if wb is not None:
        wb.save(output_dir / "export.xlsx")


# ===============================================================
# GUI – Tkinter (Fortschritt, Auswahl, Ollama, Review, Output-Tree, Config-Editor)
# ===============================================================

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1100x740")
        self.minsize(980, 640)

        self.cfg = load_config()
        self.log_queue: "queue.Queue" = queue.Queue()
        self.worker: Optional[threading.Thread] = None

        self.create_widgets()
        self.refresh_lists()
        self.after(100, self.poll_log_queue)

    # ----------------- Widgets -----------------
    def create_widgets(self):
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        # Pfade & Einstellungen
        path_frame = ttk.LabelFrame(frm, text="Pfade & Einstellungen")
        path_frame.pack(fill=tk.X)

        ttk.Label(path_frame, text="Input-Ordner:").grid(row=0, column=0, sticky=tk.W, padx=6, pady=6)
        self.var_input = tk.StringVar(value=self.cfg.get("input_dir"))
        ttk.Entry(path_frame, textvariable=self.var_input, width=60).grid(row=0, column=1, sticky=tk.W, padx=6)
        ttk.Button(path_frame, text="Wählen…", command=self.choose_input).grid(row=0, column=2, padx=6)

        ttk.Label(path_frame, text="Output-Ordner:").grid(row=1, column=0, sticky=tk.W, padx=6, pady=6)
        self.var_output = tk.StringVar(value=self.cfg.get("output_dir"))
        ttk.Entry(path_frame, textvariable=self.var_output, width=60).grid(row=1, column=1, sticky=tk.W, padx=6)
        ttk.Button(path_frame, text="Wählen…", command=self.choose_output).grid(row=1, column=2, padx=6)
        ttk.Button(path_frame, text="Ordner öffnen", command=self.open_output_folder).grid(row=1, column=3, padx=6)

        ttk.Label(path_frame, text="OCR-Seiten (max):").grid(row=2, column=0, sticky=tk.W, padx=6, pady=6)
        self.var_maxpages = tk.IntVar(value=int(self.cfg.get("max_pages_ocr", 3)))
        ttk.Spinbox(path_frame, from_=1, to=10, textvariable=self.var_maxpages, width=6).grid(row=2, column=1, sticky=tk.W, padx=6)

        ttk.Label(path_frame, text="Tesseract-Sprachen:").grid(row=2, column=2, sticky=tk.W, padx=6)
        self.var_langs = tk.StringVar(value=self.cfg.get("langs_ocr", "deu+eng"))
        ttk.Entry(path_frame, textvariable=self.var_langs, width=12).grid(row=2, column=3, sticky=tk.W, padx=6)

        self.var_use_ollama = tk.BooleanVar(value=bool(self.cfg.get("use_ollama", False)))
        ttk.Checkbutton(path_frame, text="Ollama verwenden (Lieferant normalisieren)", variable=self.var_use_ollama).grid(row=3, column=0, columnspan=2, sticky=tk.W, padx=6)

        self.var_review = tk.BooleanVar(value=bool(self.cfg.get("review_before_save", True)))
        ttk.Checkbutton(path_frame, text="Vor dem Speichern prüfen (Dialog)", variable=self.var_review).grid(row=3, column=2, columnspan=2, sticky=tk.W, padx=6)

        ttk.Button(path_frame, text="Config anzeigen", command=self.show_config_window).grid(row=0, column=3, padx=6)

        # Buttons
        btn_frame = ttk.Frame(frm)
        btn_frame.pack(fill=tk.X, pady=8)
        ttk.Button(btn_frame, text="Verarbeiten (alle)", command=self.start_processing_all).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Nur ausgewählte verarbeiten", command=self.start_processing_selected).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Liste aktualisieren", command=self.refresh_lists).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Config speichern", command=self.save_current_config).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Beenden", command=self.exit_app).pack(side=tk.RIGHT, padx=4)

        # Mittlerer Bereich: links Input-Dateien, rechts Output-Kontrollfenster
        center = ttk.Frame(frm)
        center.pack(fill=tk.BOTH, expand=True)

        left = ttk.LabelFrame(center, text="PDFs im Input-Ordner (Mehrfachauswahl möglich)")
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,8))
        self.listbox = tk.Listbox(left, selectmode=tk.EXTENDED)
        self.listbox.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        right = ttk.LabelFrame(center, text="Output-Kontrollfenster (Jahr/Lieferant/Dateien)")
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(right, columns=("typ"), show="tree")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # Protokoll + Fortschritt unten
        bottom = ttk.LabelFrame(frm, text="Protokoll & Fortschritt")
        bottom.pack(fill=tk.BOTH, expand=False)
        self.txt = tk.Text(bottom, height=10, wrap=tk.WORD)
        self.txt.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.progress = ttk.Progressbar(bottom, mode="determinate")
        self.progress.pack(fill=tk.X, padx=6, pady=(0,6))

    # ----------------- Output-Kontrollfenster befüllen -----------------
    def populate_output_tree(self):
        self.tree.delete(*self.tree.get_children())
        out = Path(self.var_output.get() or self.cfg.get("output_dir", "output"))
        out.mkdir(exist_ok=True)
        # Struktur: Jahr -> Lieferant -> Dateien
        for year_dir in sorted([p for p in out.iterdir() if p.is_dir()]):
            yid = self.tree.insert("", "end", text=year_dir.name)
            for supp_dir in sorted([p for p in year_dir.iterdir() if p.is_dir()]):
                sid = self.tree.insert(yid, "end", text=supp_dir.name)
                for f in sorted(supp_dir.glob("*.pdf")):
                    self.tree.insert(sid, "end", text=f.name)

    # ----------------- Queue/Log/Progress -----------------
    def log(self, msg: str):
        self.log_queue.put(("log", msg))

    def set_progress(self, cur: int, total: int):
        self.log_queue.put(("progress", cur, total))

    def poll_log_queue(self):
        try:
            while True:
                item = self.log_queue.get_nowait()
                kind = item[0]
                if kind == "log":
                    self.txt.insert('end', item[1] + "\n")
                    self.txt.see('end')
                elif kind == "progress":
                    cur, total = item[1], item[2]
                    self.progress.configure(maximum=total)
                    self.progress['value'] = cur
        except queue.Empty:
            pass
        self.after(100, self.poll_log_queue)

    # ----------------- Pfad-Handling -----------------
    def choose_input(self):
        d = filedialog.askdirectory(initialdir=self.var_input.get() or ".")
        if d:
            self.var_input.set(d)
            self.refresh_lists()

    def choose_output(self):
        d = filedialog.askdirectory(initialdir=self.var_output.get() or ".")
        if d:
            self.var_output.set(d)
            self.populate_output_tree()

    def open_output_folder(self):
        out = Path(self.var_output.get() or self.cfg.get("output_dir", "output"))
        out.mkdir(exist_ok=True)
        try:
            if os.name == "nt":
                os.startfile(out)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(out)])
            else:
                subprocess.Popen(["xdg-open", str(out)])
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Konnte Ordner nicht öffnen: {e}")

    # ----------------- Listen aktualisieren -----------------
    def refresh_lists(self):
        # Input-Liste
        self.listbox.delete(0, 'end')
        inp = Path(self.var_input.get() or self.cfg.get("input_dir", "input"))
        inp.mkdir(exist_ok=True)
        for p in sorted(inp.glob("*.pdf")):
            self.listbox.insert('end', str(p))
        # Output-Baum
        self.populate_output_tree()

    # ----------------- Config speichern / anzeigen -----------------
    def _prepare_cfg_from_ui(self):
        self.cfg["input_dir"] = self.var_input.get()
        self.cfg["output_dir"] = self.var_output.get()
        self.cfg["max_pages_ocr"] = int(self.var_maxpages.get())
        self.cfg["langs_ocr"] = self.var_langs.get()
        self.cfg["use_ollama"] = bool(self.var_use_ollama.get())
        self.cfg["review_before_save"] = bool(self.var_review.get())

    def save_current_config(self):
        self._prepare_cfg_from_ui()
        save_config(self.cfg)
        messagebox.showinfo(APP_NAME, "Config gespeichert.")

    def show_config_window(self):
        win = tk.Toplevel(self)
        win.title("config.json – ansehen & bearbeiten")
        win.geometry("720x520")
        win.transient(self)
        win.grab_set()

        # Textfeld + Scrollbars
        txt = tk.Text(win, wrap=tk.NONE)
        xscroll = tk.Scrollbar(win, orient=tk.HORIZONTAL, command=txt.xview)
        yscroll = tk.Scrollbar(win, orient=tk.VERTICAL, command=txt.yview)
        txt.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
        txt.grid(row=0, column=0, sticky='nsew')
        yscroll.grid(row=0, column=1, sticky='ns')
        xscroll.grid(row=1, column=0, sticky='ew')
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(0, weight=1)

        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                content = f.read()
        except FileNotFoundError:
            content = json.dumps(DEFAULT_CONFIG, indent=2, ensure_ascii=False)
        txt.insert("1.0", content)

        btns = ttk.Frame(win)
        btns.grid(row=2, column=0, columnspan=2, sticky='ew', pady=6)

        def save_changes():
            try:
                new_content = txt.get("1.0", tk.END)
                cfg = json.loads(new_content)
                save_config(cfg)
                self.cfg = cfg
                # GUI Felder aktualisieren
                self.var_input.set(self.cfg.get("input_dir", "input"))
                self.var_output.set(self.cfg.get("output_dir", "output"))
                self.var_maxpages.set(int(self.cfg.get("max_pages_ocr", 3)))
                self.var_langs.set(self.cfg.get("langs_ocr", "deu+eng"))
                self.var_use_ollama.set(bool(self.cfg.get("use_ollama", False)))
                self.var_review.set(bool(self.cfg.get("review_before_save", True)))
                messagebox.showinfo(APP_NAME, "Config gespeichert & übernommen.")
                win.destroy()
            except Exception as e:
                messagebox.showerror(APP_NAME, f"Fehler beim Speichern: {e}")

        ttk.Button(btns, text="Speichern", command=save_changes).pack(side=tk.RIGHT, padx=6)
        ttk.Button(btns, text="Schließen", command=win.destroy).pack(side=tk.RIGHT)

    # ----------------- Review-Dialog -----------------
    def review_fields_sync(self, fields: dict, pdf_path: Path) -> Optional[dict]:
        done = threading.Event()
        result: Dict[str, Any] = {"val": None}

        def open_dialog():
            win = tk.Toplevel(self)
            win.title(f"Felder prüfen – {pdf_path.name}")
            win.grab_set()
            win.transient(self)

            tk.Label(win, text="Gefundene Felder prüfen/anpassen:").pack(anchor='w', padx=10, pady=(10,4))
            frm = ttk.Frame(win)
            frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            tk.Label(frm, text="Rechnungsdatum (YYYY-MM-DD):").grid(row=0, column=0, sticky='w')
            var_date = tk.StringVar(value=fields.get("invoice_date").isoformat() if fields.get("invoice_date") else "")
            tk.Entry(frm, textvariable=var_date, width=30).grid(row=0, column=1, sticky='w')

            tk.Label(frm, text="Lieferant:").grid(row=1, column=0, sticky='w', pady=(6,0))
            var_supp = tk.StringVar(value=fields.get("supplier") or "")
            tk.Entry(frm, textvariable=var_supp, width=40).grid(row=1, column=1, sticky='w', pady=(6,0))

            tk.Label(frm, text="Rechnungsnummer:").grid(row=2, column=0, sticky='w', pady=(6,0))
            var_no = tk.StringVar(value=fields.get("invoice_no") or "")
            tk.Entry(frm, textvariable=var_no, width=30).grid(row=2, column=1, sticky='w', pady=(6,0))

            btns = ttk.Frame(win)
            btns.pack(fill=tk.X, padx=10, pady=(6,10))

            def accept():
                edited = {
                    "invoice_date": parse_date(var_date.get()) if var_date.get().strip() else None,
                    "supplier": var_supp.get().strip() or None,
                    "invoice_no": var_no.get().strip() or None,
                }
                result["val"] = edited
                win.destroy()
                done.set()

            def cancel():
                result["val"] = None
                win.destroy()
                done.set()

            ttk.Button(btns, text="OK", command=accept).pack(side=tk.LEFT)
            ttk.Button(btns, text="Abbrechen", command=cancel).pack(side=tk.RIGHT)

        self.after(0, open_dialog)
        done.wait()
        return result["val"]

    # ----------------- Worker -----------------
    def _run_worker(self, files: Optional[List[Path]]):
        self.txt.delete('1.0', 'end')
        self.progress['value'] = 0
        self.log("Starte Verarbeitung…")

        def work():
            try:
                process_all(
                    self.cfg,
                    log=self.log,
                    selected_files=[str(f) for f in files] if files else None,
                    progress=self.set_progress,
                    review_cb=self.review_fields_sync if self.var_review.get() else None,
                )
                self.log("Fertig.")
                self.after(0, self.populate_output_tree)
            except Exception as e:
                self.log(f"Fehler: {e}")
        self.worker = threading.Thread(target=work, daemon=True)
        self.worker.start()

    def start_processing_all(self):
        if self.worker and self.worker.is_alive():
            messagebox.showinfo(APP_NAME, "Verarbeitung läuft bereits…")
            return
        self._prepare_cfg_from_ui()
        self._run_worker(files=None)

    def start_processing_selected(self):
        if self.worker and self.worker.is_alive():
            messagebox.showinfo(APP_NAME, "Verarbeitung läuft bereits…")
            return
        self._prepare_cfg_from_ui()
        sels = [self.listbox.get(i) for i in self.listbox.curselection()]
        if not sels:
            messagebox.showinfo(APP_NAME, "Bitte eine oder mehrere PDF-Dateien in der Liste markieren.")
            return
        files = [Path(s) for s in sels]
        self._run_worker(files=files)

    # ----------------- Exit -----------------
    def exit_app(self):
        self.destroy()


if __name__ == "__main__":
    try:
        App().mainloop()
    except KeyboardInterrupt:
        pass
